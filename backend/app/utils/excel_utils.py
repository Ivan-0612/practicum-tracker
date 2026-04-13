import os
import re
from io import BytesIO
from typing import Any, Dict, Tuple
from urllib.error import HTTPError, URLError
from urllib.parse import quote
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries


def _normalize_supabase_http_url(raw_url: str) -> str:
    value = (raw_url or "").strip().strip('"').strip("'")
    if not value:
        return ""

    if value.startswith("http://") or value.startswith("https://"):
        return value.rstrip("/")

    if value.startswith("postgresql://") or value.startswith("postgres://"):
        parsed = urlparse(value)
        # Ejemplo user: postgres.<project_ref>
        username = parsed.username or ""
        if "." in username:
            maybe_ref = username.split(".", 1)[1]
            if maybe_ref:
                return f"https://{maybe_ref}.supabase.co"

    return ""


def _build_storage_download_url(storage_path: str) -> str:
    supabase_url = _normalize_supabase_http_url(os.getenv("SUPABASE_URL") or "")
    bucket = os.getenv("SUPABASE_STORAGE_BUCKET", "plantillas")
    if not supabase_url:
        raise ValueError(
            "SUPABASE_URL inválida. Debe ser https://<project-ref>.supabase.co"
        )

    encoded_path = quote(storage_path.lstrip("/"))
    return f"{supabase_url}/storage/v1/object/{bucket}/{encoded_path}"


def _build_storage_upload_url(storage_path: str) -> str:
    return _build_storage_download_url(storage_path)


def _download_template_bytes(storage_path: str) -> bytes:
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not service_key:
        raise ValueError("Falta SUPABASE_SERVICE_ROLE_KEY en variables de entorno")

    url = _build_storage_download_url(storage_path)
    request = Request(
        url,
        headers={
            "Authorization": f"Bearer {service_key}",
            "apikey": service_key,
        },
        method="GET",
    )

    try:
        with urlopen(request, timeout=30) as response:
            return response.read()
    except HTTPError as exc:
        raise ValueError(f"No se pudo descargar la plantilla XLSX ({exc.code})") from exc
    except URLError as exc:
        raise ValueError("No se pudo conectar con Supabase Storage") from exc


def subir_plantilla_excel(storage_path: str, content: bytes):
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not service_key:
        raise ValueError("Falta SUPABASE_SERVICE_ROLE_KEY en variables de entorno")

    url = _build_storage_upload_url(storage_path)
    request = Request(
        url,
        data=content,
        headers={
            "Authorization": f"Bearer {service_key}",
            "apikey": service_key,
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "x-upsert": "true",
        },
        method="POST",
    )
    try:
        with urlopen(request, timeout=30):
            return
    except HTTPError as exc:
        raise ValueError(f"No se pudo subir la plantilla XLSX ({exc.code})") from exc
    except URLError as exc:
        raise ValueError(f"No se pudo conectar con Supabase Storage: {exc.reason}") from exc


def borrar_plantilla_excel(storage_path: str):
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not service_key:
        raise ValueError("Falta SUPABASE_SERVICE_ROLE_KEY en variables de entorno")

    url = _build_storage_download_url(storage_path)
    request = Request(
        url,
        headers={
            "Authorization": f"Bearer {service_key}",
            "apikey": service_key,
        },
        method="DELETE",
    )
    try:
        with urlopen(request, timeout=30):
            return
    except HTTPError as exc:
        # 404 se considera idempotente para borrado.
        if exc.code == 404:
            return
        raise ValueError(f"No se pudo borrar la plantilla XLSX ({exc.code})") from exc
    except URLError as exc:
        raise ValueError(f"No se pudo conectar con Supabase Storage: {exc.reason}") from exc


def _resolve_target(mapping_entry: Any) -> Tuple[str, str]:
    if isinstance(mapping_entry, str):
        if "!" in mapping_entry:
            sheet, cell = mapping_entry.split("!", 1)
            return sheet, cell
        return "EVALUACION", mapping_entry

    if isinstance(mapping_entry, dict):
        sheet = mapping_entry.get("sheet") or "EVALUACION"
        cell = mapping_entry.get("cell")
        if not cell:
            raise ValueError("Cada mapeo debe incluir la celda destino")
        return sheet, cell

    raise ValueError("Formato de mapeo no válido")


def _find_sheet_name_case_insensitive(wb, expected_name: str) -> str:
    expected = (expected_name or "").strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == expected:
            return name
    raise ValueError(f"La plantilla no contiene la hoja '{expected_name}'")


def _to_grade_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value)


def _split_args(expr: str):
    args = []
    current = []
    depth = 0
    in_quotes = False
    quote_char = ""

    for ch in expr:
        if ch in ('"', "'"):
            if in_quotes and ch == quote_char:
                in_quotes = False
                quote_char = ""
            elif not in_quotes:
                in_quotes = True
                quote_char = ch
            current.append(ch)
            continue

        if not in_quotes:
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
            elif ch == ',' and depth == 0:
                args.append(''.join(current).strip())
                current = []
                continue

        current.append(ch)

    if current:
        args.append(''.join(current).strip())

    return args


def _split_condition(expr: str):
    ops = ["<=", ">=", "<>", "<", ">", "="]
    depth = 0
    in_quotes = False
    quote_char = ""
    i = 0

    while i < len(expr):
        ch = expr[i]
        if ch in ('"', "'"):
            if in_quotes and ch == quote_char:
                in_quotes = False
                quote_char = ""
            elif not in_quotes:
                in_quotes = True
                quote_char = ch
            i += 1
            continue

        if not in_quotes:
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
            elif depth == 0:
                for op in ops:
                    if expr.startswith(op, i):
                        left = expr[:i].strip()
                        right = expr[i + len(op) :].strip()
                        return left, op, right
        i += 1

    return None, None, None


def _to_number(value: Any):
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, str):
        text = value.strip().replace(',', '.')
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _numeric(value: Any) -> float:
    n = _to_number(value)
    return n if n is not None else 0.0


def _resolve_sheet_and_cell(expr: str, current_sheet_name: str):
    token = expr.strip()
    if '!' in token:
        sheet_name, cell = token.split('!', 1)
        return sheet_name.strip("'\" "), cell.strip()
    return current_sheet_name, token


def _resolve_sheet_and_range(expr: str, current_sheet_name: str):
    token = expr.strip()
    if '!' in token:
        sheet_name, rng = token.split('!', 1)
        return sheet_name.strip("'\" "), rng.strip()
    return current_sheet_name, token


def _evaluate_range_numeric_sum(
    wb,
    sheet_name: str,
    range_expr: str,
    cache: Dict[str, Any],
    stack: set,
):
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_expr)
    except Exception:
        return 0.0

    total = 0.0
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell_ref = f"{get_column_letter(col)}{row}"
            val = _evaluate_cell(wb, sheet_name, cell_ref, cache, stack)
            total += _numeric(val)
    return total


def _extract_row_from_cell_ref(cell_ref: str) -> int:
    m = re.search(r"(\d+)$", (cell_ref or "").strip())
    if not m:
        raise ValueError(f"Referencia de celda inválida en mapping: '{cell_ref}'")
    return int(m.group(1))


def _evaluate_condition(expr: str, wb, current_sheet_name: str, cache: Dict[str, Any], stack: set):
    left, op, right = _split_condition(expr)
    if not op:
        return bool(_evaluate_expression(expr, wb, current_sheet_name, cache, stack))

    lval = _evaluate_expression(left, wb, current_sheet_name, cache, stack)
    rval = _evaluate_expression(right, wb, current_sheet_name, cache, stack)

    lnum = _to_number(lval)
    rnum = _to_number(rval)
    if lnum is not None and rnum is not None:
        a, b = lnum, rnum
    else:
        a, b = str(lval), str(rval)

    if op == "<":
        return a < b
    if op == ">":
        return a > b
    if op == "<=":
        return a <= b
    if op == ">=":
        return a >= b
    if op == "=":
        return a == b
    if op == "<>":
        return a != b
    return False


def _evaluate_cell(wb, sheet_name: str, cell_ref: str, cache: Dict[str, Any], stack: set):
    key = f"{sheet_name}!{cell_ref}"
    if key in cache:
        return cache[key]
    if key in stack:
        return 0

    stack.add(key)
    sheet_real = _find_sheet_name_case_insensitive(wb, sheet_name)
    value = wb[sheet_real][cell_ref].value

    if isinstance(value, str) and value.startswith('='):
        result = _evaluate_expression(value[1:], wb, sheet_real, cache, stack)
    else:
        result = value

    cache[key] = result
    stack.remove(key)
    return result


def _evaluate_expression(expr: str, wb, current_sheet_name: str, cache: Dict[str, Any], stack: set):
    raw = expr.strip()
    if not raw:
        return 0

    if (raw.startswith('"') and raw.endswith('"')) or (raw.startswith("'") and raw.endswith("'")):
        return raw[1:-1]

    upper = raw.upper()
    if upper.startswith("SUM(") and raw.endswith(")"):
        args = _split_args(raw[4:-1])
        total = 0.0
        for arg in args:
            arg_str = arg.strip()
            if ':' in arg_str:
                sheet_name, range_expr = _resolve_sheet_and_range(arg_str, current_sheet_name)
                total += _evaluate_range_numeric_sum(wb, sheet_name, range_expr, cache, stack)
            else:
                total += _numeric(_evaluate_expression(arg_str, wb, current_sheet_name, cache, stack))
        return total

    if upper.startswith("PRODUCT(") and raw.endswith(")"):
        args = _split_args(raw[8:-1])
        result = 1.0
        for arg in args:
            result *= _numeric(_evaluate_expression(arg, wb, current_sheet_name, cache, stack))
        return result

    if upper.startswith("IF(") and raw.endswith(")"):
        args = _split_args(raw[3:-1])
        if len(args) != 3:
            return 0
        cond = _evaluate_condition(args[0], wb, current_sheet_name, cache, stack)
        return _evaluate_expression(args[1] if cond else args[2], wb, current_sheet_name, cache, stack)

    cell_match = re.fullmatch(r"(?:(?:'[^']+'|[A-Za-z0-9_ ]+)!)?[A-Z]{1,3}[0-9]{1,7}", raw, flags=re.IGNORECASE)
    if cell_match:
        sheet_name, cell_ref = _resolve_sheet_and_cell(raw, current_sheet_name)
        return _evaluate_cell(wb, sheet_name, cell_ref.upper(), cache, stack)

    # Reemplazamos referencias de celda por su valor numérico para evaluar operaciones básicas.
    ref_pattern = re.compile(r"(?:(?:'[^']+'|[A-Za-z0-9_ ]+)!)?[A-Z]{1,3}[0-9]{1,7}", re.IGNORECASE)

    def repl(match):
        token = match.group(0)
        sheet_name, cell_ref = _resolve_sheet_and_cell(token, current_sheet_name)
        value = _evaluate_cell(wb, sheet_name, cell_ref.upper(), cache, stack)
        return str(_numeric(value))

    py_expr = ref_pattern.sub(repl, raw)
    safe_expr = py_expr.replace("^", "**")
    if not re.fullmatch(r"[0-9\.\+\-\*/\(\)\s]+", safe_expr):
        return 0

    try:
        return eval(safe_expr, {"__builtins__": {}}, {})
    except Exception:
        return 0


def generar_excel_evaluado(especialidad, respuestas_json: Dict[str, Any], mapping_json: Dict[str, Any]) -> Tuple[bytes, str]:
    storage_path = especialidad.plantilla_excel_storage_path
    mapping = mapping_json or {}

    if not storage_path:
        raise ValueError("La especialidad no tiene plantilla Excel configurada")

    if not isinstance(mapping, dict) or not mapping:
        raise ValueError("No hay mapeo Excel global configurado")

    # Evita resultados erróneos (p.ej. nota 0) cuando el mapping usa claves
    # que no corresponden a los IDs de unidades realmente evaluados.
    respuestas_ids = {k for k, v in (respuestas_json or {}).items() if isinstance(v, dict)}
    mapping_ids = set(mapping.keys())
    if respuestas_ids and mapping_ids.isdisjoint(respuestas_ids):
        raise ValueError(
            "El mapping global no coincide con los IDs de unidades de competencia evaluadas. "
            "Revisa y vuelve a guardar el mapping usando las claves correctas (ej: a1_01, a2_03, ...)."
        )

    template_bytes = _download_template_bytes(storage_path)
    wb = load_workbook(BytesIO(template_bytes), data_only=False)

    # Solo volcamos Unidades de Competencia (valor_nivel); NIC queda fuera.
    # Regla solicitada: marcar exclusivamente columnas B/C/D con "X" según nivel.
    for elemento_id, mapping_entry in mapping.items():
        respuesta = respuestas_json.get(elemento_id)
        if not isinstance(respuesta, dict):
            continue

        valor_nivel = respuesta.get("valor_nivel")
        if valor_nivel is None:
            continue

        sheet_name, cell = _resolve_target(mapping_entry)
        sheet_name_real = _find_sheet_name_case_insensitive(wb, sheet_name)
        row = _extract_row_from_cell_ref(cell)

        ws_eval = wb[sheet_name_real]
        ws_eval[f"B{row}"] = ""
        ws_eval[f"C{row}"] = ""
        ws_eval[f"D{row}"] = ""

        if valor_nivel == 1:
            ws_eval[f"B{row}"] = "X"
        elif valor_nivel == 2:
            ws_eval[f"C{row}"] = "X"
        elif valor_nivel == 3:
            ws_eval[f"D{row}"] = "X"
        else:
            raise ValueError(
                f"valor_nivel inválido para {elemento_id}: {valor_nivel}. Debe ser 1, 2 o 3"
            )

    hoja2_real = _find_sheet_name_case_insensitive(wb, "hoja2")

    final_grade_value = wb[hoja2_real]["B86"].value
    if isinstance(final_grade_value, str) and final_grade_value.startswith('='):
        final_grade_value = _evaluate_expression(
            final_grade_value[1:],
            wb,
            hoja2_real,
            cache={},
            stack=set(),
        )
    final_grade_text = _to_grade_text(final_grade_value)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue(), final_grade_text
