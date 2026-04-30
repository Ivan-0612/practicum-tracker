from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy.orm import Session
from sqlalchemy import or_
from ..database import get_db
from .. import models, schemas, security
import secrets
from datetime import date, datetime
from io import BytesIO
from pydantic import ValidationError
from openpyxl import load_workbook
from uuid import UUID
import re
import unicodedata
from ..utils.periodo_academico_utils import normalizar_periodo_academico

router = APIRouter(prefix="/api/v1/alumnos", tags=["Alumnos"])


def _normalizar_texto(valor):
    if valor is None:
        return ""
    texto = str(valor).strip().lower().replace("_", " ").replace("-", " ")
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", texto)


def _extraer_numero(valor):
    if valor is None:
        return None
    if isinstance(valor, bool):
        return None
    if isinstance(valor, int):
        return valor
    if isinstance(valor, float):
        return int(valor) if valor.is_integer() else None
    match = re.search(r"\d+", str(valor))
    return int(match.group()) if match else None


def _obtener_valor_fila(row, header_map, aliases):
    for alias in aliases:
        idx = header_map.get(_normalizar_texto(alias))
        if idx is not None and idx < len(row):
            value = row[idx]
            if value is not None and str(value).strip() != "":
                return value
    return None


def _resolver_especialidad(db: Session, valor):
    if valor is None:
        return None

    texto = str(valor).strip()
    if not texto:
        return None

    try:
        especialidad_uuid = UUID(texto)
        especialidad = (
            db.query(models.Especialidad)
            .filter(models.Especialidad.id == especialidad_uuid)
            .first()
        )
        if especialidad:
            return especialidad
    except Exception:
        pass

    texto_normalizado = _normalizar_texto(texto)
    if not texto_normalizado:
        return None

    especialidades = db.query(models.Especialidad).all()
    for especialidad in especialidades:
        if _normalizar_texto(especialidad.nombre) == texto_normalizado:
            return especialidad

    return None


def _crear_alumno_desde_datos(db: Session, alumno_in: schemas.AlumnoCreate):
    periodo_academico = normalizar_periodo_academico(alumno_in.periodo_academico)

    existe_usuario = (
        db.query(models.Usuario)
        .filter(models.Usuario.email == alumno_in.email_acceso)
        .first()
    )
    if existe_usuario:
        raise HTTPException(
            status_code=400,
            detail="Ya existe un estudiante registrado con este email.",
        )

    tutor_hospital = (
        db.query(models.Usuario)
        .filter(
            models.Usuario.email == alumno_in.email_tutor_hospital,
            models.Usuario.rol == "profesor",
            models.Usuario.activo == True,
        )
        .first()
    )
    if not tutor_hospital:
        raise HTTPException(
            status_code=404,
            detail=f"No existe el Tutor de Hospital: {alumno_in.email_tutor_hospital}",
        )

    tutor_universidad = (
        db.query(models.Usuario)
        .filter(
            models.Usuario.email == alumno_in.email_tutor_universidad,
            models.Usuario.rol == "profesor",
            models.Usuario.activo == True,
        )
        .first()
    )
    if not tutor_universidad:
        raise HTTPException(
            status_code=404,
            detail=f"No existe el Tutor de Universidad: {alumno_in.email_tutor_universidad}",
        )

    especialidad = (
        db.query(models.Especialidad)
        .filter(models.Especialidad.id == alumno_in.especialidad_id)
        .first()
    )
    if not especialidad:
        raise HTTPException(
            status_code=404, detail="La especialidad seleccionada no existe"
        )

    nuevo_usuario = models.Usuario(
        email=alumno_in.email_acceso,
        password_hash=security.get_password_hash(alumno_in.password_acceso),
        rol="estudiante",
    )
    db.add(nuevo_usuario)
    db.flush()

    nuevo_alumno = models.Alumno(
        usuario_id=nuevo_usuario.id,
        curso=alumno_in.curso,
        grupo=alumno_in.grupo,
        numero_rotacion=alumno_in.numero_rotacion,
        codigo_anonimo=f"ALU-{secrets.token_hex(3).upper()}",
        nombre_cifrado=security.cifrar_dato(alumno_in.nombre),
        apellidos_cifrado=security.cifrar_dato(alumno_in.apellidos),
        email_cifrado=security.cifrar_dato(alumno_in.email_personal),
    )
    db.add(nuevo_alumno)
    db.flush()

    nueva_rotacion = models.Rotacion(
        alumno_id=nuevo_alumno.id,
        especialidad_id=especialidad.id,
        curso=alumno_in.curso,
        numero_rotacion=alumno_in.numero_rotacion,
        periodo_academico=periodo_academico,
        centro_practicas=alumno_in.centro_practicas,
    )
    db.add(nueva_rotacion)
    db.flush()

    asignacion_hosp = models.AsignacionTutor(
        tutor_id=tutor_hospital.id, rotacion_id=nueva_rotacion.id, tipo_tutor="hospital"
    )
    asignacion_uni = models.AsignacionTutor(
        tutor_id=tutor_universidad.id,
        rotacion_id=nueva_rotacion.id,
        tipo_tutor="universidad",
    )
    db.add(asignacion_hosp)
    db.add(asignacion_uni)

    try:
        db.commit()
    except Exception:
        db.rollback()
        raise

    db.refresh(nuevo_alumno)
    return nuevo_alumno


def _obtener_valor_requerido(fila, header_map, aliases, nombre_campo):
    valor = _obtener_valor_fila(fila, header_map, aliases)
    if valor is None:
        raise ValueError(f"Falta la columna '{nombre_campo}' o está vacía")
    return valor


def _crear_rotacion_alumno(
    db: Session,
    alumno,
    curso: int,
    numero_rotacion: int,
    periodo_academico: str,
    centro_practicas: str,
    especialidad,
    email_tutor_hospital: str,
    email_tutor_universidad: str,
):
    periodo_academico = normalizar_periodo_academico(periodo_academico)

    existe_rotacion_curso = (
        db.query(models.Rotacion)
        .filter(
            models.Rotacion.alumno_id == alumno.id,
            models.Rotacion.numero_rotacion == numero_rotacion,
            models.Rotacion.curso == curso,
        )
        .first()
    )
    if existe_rotacion_curso:
        raise HTTPException(
            status_code=400,
            detail=(
                f"El alumno ya tiene asignada la Rotación {numero_rotacion} "
                f"para {curso}º Curso."
            ),
        )

    tutor_hospital = (
        db.query(models.Usuario)
        .filter(
            models.Usuario.email == email_tutor_hospital,
            models.Usuario.rol == "profesor",
            models.Usuario.activo == True,
        )
        .first()
    )
    if not tutor_hospital:
        raise HTTPException(
            status_code=404, detail="Tutor de Hospital no válido o no existe"
        )

    tutor_universidad = (
        db.query(models.Usuario)
        .filter(
            models.Usuario.email == email_tutor_universidad,
            models.Usuario.rol == "profesor",
            models.Usuario.activo == True,
        )
        .first()
    )
    if not tutor_universidad:
        raise HTTPException(
            status_code=404, detail="Tutor de Universidad no válido o no existe"
        )

    nueva_rotacion = models.Rotacion(
        alumno_id=alumno.id,
        especialidad_id=especialidad.id,
        curso=curso,
        numero_rotacion=numero_rotacion,
        periodo_academico=periodo_academico,
        centro_practicas=centro_practicas,
    )
    db.add(nueva_rotacion)
    db.flush()

    asignacion_hosp = models.AsignacionTutor(
        tutor_id=tutor_hospital.id,
        rotacion_id=nueva_rotacion.id,
        tipo_tutor="hospital",
    )
    asignacion_uni = models.AsignacionTutor(
        tutor_id=tutor_universidad.id,
        rotacion_id=nueva_rotacion.id,
        tipo_tutor="universidad",
    )
    db.add(asignacion_hosp)
    db.add(asignacion_uni)

    try:
        db.commit()
    except Exception:
        db.rollback()
        raise

    db.refresh(nueva_rotacion)
    return nueva_rotacion


def _obtener_centro_valido(db: Session, centro_practicas_id: UUID):
    centro = (
        db.query(models.CentroPracticas)
        .filter(
            models.CentroPracticas.id == centro_practicas_id,
            models.CentroPracticas.activo == True,
        )
        .first()
    )
    if not centro:
        raise HTTPException(status_code=404, detail="Centro de prácticas no encontrado")

    if not centro.tutor_hospital_id or not centro.tutor_universidad_id:
        raise HTTPException(
            status_code=400,
            detail="El centro no tiene tutores configurados",
        )

    return centro


def _crear_rotacion_desde_centro(
    db: Session,
    alumno,
    curso: int,
    numero_rotacion: int,
    periodo_academico: str,
    especialidad_id: UUID,
    centro,
):
    periodo_academico = normalizar_periodo_academico(periodo_academico)

    existe_rotacion_curso = (
        db.query(models.Rotacion)
        .filter(
            models.Rotacion.alumno_id == alumno.id,
            models.Rotacion.numero_rotacion == numero_rotacion,
            models.Rotacion.curso == curso,
        )
        .first()
    )
    if existe_rotacion_curso:
        raise HTTPException(
            status_code=400,
            detail=(
                f"El alumno ya tiene asignada la Rotación {numero_rotacion} "
                f"para {curso}º Curso."
            ),
        )

    nueva_rotacion = models.Rotacion(
        alumno_id=alumno.id,
        especialidad_id=especialidad_id,
        curso=curso,
        numero_rotacion=numero_rotacion,
        periodo_academico=periodo_academico,
        centro_practicas=centro.nombre,
    )
    db.add(nueva_rotacion)
    db.flush()

    db.add(
        models.AsignacionTutor(
            tutor_id=centro.tutor_hospital_id,
            rotacion_id=nueva_rotacion.id,
            tipo_tutor="hospital",
        )
    )
    db.add(
        models.AsignacionTutor(
            tutor_id=centro.tutor_universidad_id,
            rotacion_id=nueva_rotacion.id,
            tipo_tutor="universidad",
        )
    )

    db.commit()
    db.refresh(nueva_rotacion)
    return nueva_rotacion


@router.post("/pre-registro")
def pre_registrar_alumno(
    datos: schemas.AlumnoPreRegistroCreate,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(security.get_current_user),
):
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")

    existe = (
        db.query(models.Usuario)
        .filter(models.Usuario.email == datos.email)
        .first()
    )
    if existe:
        raise HTTPException(status_code=400, detail="El correo ya está registrado")

    nuevo_usuario = models.Usuario(
        email=datos.email,
        password_hash=security.get_password_hash(secrets.token_urlsafe(16)),
        rol="estudiante",
        registro_completado=False,
        curso_pendiente=datos.curso,
        activo=True,
    )
    db.add(nuevo_usuario)
    db.commit()

    return {"mensaje": "Alumno pre-registrado correctamente", "email": datos.email}


@router.post("/", response_model=schemas.AlumnoResponse)
def crear_alumno(alumno_in: schemas.AlumnoCreate, db: Session = Depends(get_db)):
    return _crear_alumno_desde_datos(db, alumno_in)


@router.post("/importar-excel", response_model=schemas.ImportAlumnosResponse)
def importar_alumnos_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(security.get_current_user),
):
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser XLSX")

    try:
        workbook = load_workbook(BytesIO(file.file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="El Excel no se pudo leer")

    if not workbook.sheetnames:
        raise HTTPException(status_code=400, detail="El Excel no contiene hojas válidas")

    worksheet = workbook[workbook.sheetnames[0]]
    filas = list(worksheet.iter_rows(values_only=True))
    if len(filas) < 2:
        raise HTTPException(status_code=400, detail="El Excel no contiene filas de datos")

    cabeceras = filas[0]
    header_map = {}
    for idx, cabecera in enumerate(cabeceras):
        if cabecera is None:
            continue
        clave = _normalizar_texto(cabecera)
        if clave:
            header_map[clave] = idx

    aliases_email = [
        "email",
        "correo",
        "correo electronico",
        "email alumno",
        "correo alumno",
    ]
    idx_email = None
    for alias in aliases_email:
        idx = header_map.get(_normalizar_texto(alias))
        if idx is not None:
            idx_email = idx
            break

    if idx_email is None:
        raise HTTPException(
            status_code=400,
            detail="El Excel debe incluir una columna de correo (email/correo)",
        )

    total_filas = 0
    creados = []
    fallos = []
    seen_emails = set()

    for fila_numero, fila in enumerate(filas[1:], start=2):
        if not any(valor is not None and str(valor).strip() for valor in fila):
            continue

        total_filas += 1
        email_acceso = None

        try:
            if idx_email >= len(fila) or fila[idx_email] is None:
                raise ValueError("Correo vacío")

            email_acceso = str(fila[idx_email]).strip()
            if not email_acceso:
                raise ValueError("Correo vacío")

            email_normalizado = email_acceso.lower()
            if email_normalizado in seen_emails:
                fallos.append(
                    {
                        "fila": fila_numero,
                        "email_acceso": email_acceso,
                        "motivo": "Duplicado dentro del mismo Excel",
                    }
                )
                continue

            if (
                db.query(models.Usuario)
                .filter(models.Usuario.email == email_acceso)
                .first()
            ):
                fallos.append(
                    {
                        "fila": fila_numero,
                        "email_acceso": email_acceso,
                        "motivo": "Ya existe un estudiante registrado con este email",
                    }
                )
                continue

            nuevo_usuario = models.Usuario(
                email=email_acceso,
                password_hash=security.get_password_hash(secrets.token_urlsafe(16)),
                rol="estudiante",
                registro_completado=False,
                activo=True,
            )
            db.add(nuevo_usuario)
            db.commit()
            seen_emails.add(email_normalizado)
            creados.append(
                {
                    "fila": fila_numero,
                    "email_acceso": email_acceso,
                    "alumno_id": str(nuevo_usuario.id),
                }
            )

        except ValidationError:
            db.rollback()
            fallos.append(
                {
                    "fila": fila_numero,
                    "email_acceso": email_acceso,
                    "motivo": "Correo inválido",
                }
            )
        except ValueError as exc:
            db.rollback()
            fallos.append(
                {
                    "fila": fila_numero,
                    "email_acceso": email_acceso,
                    "motivo": str(exc),
                }
            )
        except Exception:
            db.rollback()
            fallos.append(
                {
                    "fila": fila_numero,
                    "email_acceso": email_acceso,
                    "motivo": "Error inesperado al procesar la fila",
                }
            )

    duplicados = len([f for f in fallos if "duplicado" in f["motivo"].lower()])
    return {
        "total_filas": total_filas,
        "creados": len(creados),
        "duplicados": duplicados,
        "fallidos": len(fallos),
        "creados_detalle": creados,
        "fallos": fallos,
    }


@router.get("/mi-perfil-evaluacion")
def obtener_mi_evaluacion(
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(security.get_current_user),
):
    if current_user.rol != "estudiante":
        raise HTTPException(status_code=403, detail="Solo para alumnos")

    alumno = (
        db.query(models.Alumno)
        .filter(models.Alumno.usuario_id == current_user.id)
        .first()
    )
    rotaciones = (
        db.query(models.Rotacion)
        .filter(models.Rotacion.alumno_id == alumno.id)
        .order_by(models.Rotacion.curso, models.Rotacion.numero_rotacion)
        .all()
    )

    hoy = date.today()
    rotaciones_data = []

    for r in rotaciones:
        asignaciones = (
            db.query(models.AsignacionTutor)
            .filter(models.AsignacionTutor.rotacion_id == r.id)
            .all()
        )

        # Mapeamos los tutores
        tutores_dict = {"hospital": "", "universidad": "", "campo": ""}
        for asig in asignaciones:
            if asig.tipo_tutor == "hospital":
                tutores_dict["hospital"] = asig.tutor.email
            elif asig.tipo_tutor == "universidad":
                tutores_dict["universidad"] = asig.tutor.email
            elif asig.tipo_tutor == "campo":
                tutores_dict["campo"] = asig.tutor.email

        fichaje_hoy = (
            db.query(models.RegistroAsistencia)
            .filter(
                models.RegistroAsistencia.rotacion_id == r.id,
                models.RegistroAsistencia.fecha == hoy,
            )
            .first()
        )

        nombre_especialidad = (
            r.especialidad.nombre if r.especialidad else "Sin especialidad"
        )

        rotaciones_data.append(
            {
                "id": str(r.id),
                "curso": r.curso if r.curso else alumno.curso,
                "numero": (
                    r.numero_rotacion if r.numero_rotacion else alumno.numero_rotacion
                ),
                "especialidad": nombre_especialidad,
                "centro_practicas": r.centro_practicas,
                "completada": r.completada,
                "tutores": tutores_dict,  # <-- AHORA ES UN DICCIONARIO
                "periodo_academico": normalizar_periodo_academico(r.periodo_academico),
            }
        )

    return {
        "alumno": {
            "nombre": security.descifrar_dato(alumno.nombre_cifrado),
            "apellidos": security.descifrar_dato(alumno.apellidos_cifrado),
        },
        "rotaciones": rotaciones_data,
    }


@router.get("/")
def listar_alumnos_por_email(
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(security.get_current_user),
):
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")

    estudiantes = (
        db.query(models.Usuario, models.Alumno)
        .join(models.Alumno, models.Alumno.usuario_id == models.Usuario.id)
        .filter(models.Usuario.rol == "estudiante")
        .all()
    )

    resultado = []
    for usuario, alumno in estudiantes:
        rotaciones_db = (
            db.query(models.Rotacion)
            .filter(models.Rotacion.alumno_id == alumno.id)
            .order_by(models.Rotacion.curso, models.Rotacion.numero_rotacion)
            .all()
        )

        lista_rotaciones = []
        for rot in rotaciones_db:
            asignaciones = (
                db.query(models.AsignacionTutor)
                .filter(models.AsignacionTutor.rotacion_id == rot.id)
                .all()
            )

            tutores_dict = {"hospital": "", "universidad": "", "campo": ""}
            for asig in asignaciones:
                if asig.tipo_tutor == "hospital":
                    tutores_dict["hospital"] = asig.tutor.email
                elif asig.tipo_tutor == "universidad":
                    tutores_dict["universidad"] = asig.tutor.email
                elif asig.tipo_tutor == "campo":
                    tutores_dict["campo"] = asig.tutor.email

            lista_rotaciones.append(
                {
                    "id": str(rot.id),
                    "curso": rot.curso or alumno.curso,
                    "numero_rotacion": rot.numero_rotacion or alumno.numero_rotacion,
                    "especialidad": (
                        rot.especialidad.nombre
                        if rot.especialidad
                        else "Sin especialidad"
                    ),
                    "centro_practicas": rot.centro_practicas,
                    "tutores": tutores_dict,  # <-- AHORA ES UN DICCIONARIO
                    "periodo_academico": normalizar_periodo_academico(rot.periodo_academico),
                    "completada": rot.completada,
                    "hospital_finalize_count": rot.hospital_finalize_count or 0,
                }
            )

        resultado.append(
            {
                "id": str(alumno.id),
                "email": usuario.email,
                "curso_actual": alumno.curso,
                "grupo": alumno.grupo,
                "rotaciones": lista_rotaciones,
            }
        )

    return resultado


@router.post("/asignar-rotacion")
def asignar_rotacion_adicional(
    datos: schemas.RotacionCreate,
    db: Session = Depends(get_db),
):
    alumno = db.query(models.Alumno).filter(models.Alumno.id == datos.alumno_id).first()
    if not alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")

    especialidad = (
        db.query(models.Especialidad)
        .filter(models.Especialidad.id == datos.especialidad_id)
        .first()
    )
    if not especialidad:
        raise HTTPException(status_code=404, detail="Especialidad no encontrada")

    _crear_rotacion_alumno(
        db=db,
        alumno=alumno,
        curso=datos.curso,
        numero_rotacion=datos.numero_rotacion,
        periodo_academico=datos.periodo_academico,
        centro_practicas=datos.centro_practicas,
        especialidad=especialidad,
        email_tutor_hospital=datos.email_tutor_hospital,
        email_tutor_universidad=datos.email_tutor_universidad,
    )

    return {"mensaje": "Rotación asignada correctamente"}


@router.post("/asignar-rotacion-automatica")
def asignar_rotacion_automatica(
    alumno_id: UUID,
    datos: schemas.RotacionAutomaticaCreate,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(security.get_current_user),
):
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")

    alumno = db.query(models.Alumno).filter(models.Alumno.id == alumno_id).first()
    if not alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")

    especialidad = (
        db.query(models.Especialidad)
        .filter(models.Especialidad.id == datos.especialidad_id)
        .first()
    )
    if not especialidad:
        raise HTTPException(status_code=404, detail="Especialidad no encontrada")

    centro = _obtener_centro_valido(db, datos.centro_practicas_id)
    _crear_rotacion_desde_centro(
        db=db,
        alumno=alumno,
        curso=datos.curso,
        numero_rotacion=datos.numero_rotacion,
        periodo_academico=datos.periodo_academico,
        especialidad_id=datos.especialidad_id,
        centro=centro,
    )

    return {"mensaje": "Rotación asignada correctamente"}


@router.get("/opciones-rotacion")
def obtener_opciones_rotacion(
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(security.get_current_user),
):
    if current_user.rol not in ["admin", "estudiante"]:
        raise HTTPException(status_code=403, detail="No autorizado")

    especialidades = db.query(models.Especialidad).order_by(models.Especialidad.nombre).all()
    centros = (
        db.query(models.CentroPracticas)
        .filter(models.CentroPracticas.activo == True)
        .order_by(models.CentroPracticas.nombre)
        .all()
    )

    return {
        "especialidades": [
            {"id": str(esp.id), "nombre": esp.nombre}
            for esp in especialidades
        ],
        "centros": [
            {
                "id": str(c.id),
                "nombre": c.nombre,
                "tutor_hospital_email": c.tutor_hospital.email if c.tutor_hospital else "",
                "tutor_universidad_email": c.tutor_universidad.email if c.tutor_universidad else "",
            }
            for c in centros
        ],
    }


@router.post("/solicitar-siguiente-rotacion")
def solicitar_siguiente_rotacion(
    datos: schemas.RotacionAutomaticaCreate,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(security.get_current_user),
):
    if current_user.rol != "estudiante":
        raise HTTPException(status_code=403, detail="Solo alumnos")

    alumno = (
        db.query(models.Alumno)
        .filter(models.Alumno.usuario_id == current_user.id)
        .first()
    )
    if not alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")

    tiene_activas = (
        db.query(models.Rotacion)
        .filter(
            models.Rotacion.alumno_id == alumno.id,
            models.Rotacion.completada == False,
        )
        .first()
    )
    if tiene_activas:
        raise HTTPException(
            status_code=400,
            detail="Primero debes completar tu rotación activa actual",
        )

    especialidad = (
        db.query(models.Especialidad)
        .filter(models.Especialidad.id == datos.especialidad_id)
        .first()
    )
    if not especialidad:
        raise HTTPException(status_code=404, detail="Especialidad no encontrada")

    centro = _obtener_centro_valido(db, datos.centro_practicas_id)
    nueva_rotacion = _crear_rotacion_desde_centro(
        db=db,
        alumno=alumno,
        curso=datos.curso,
        numero_rotacion=datos.numero_rotacion,
        periodo_academico=datos.periodo_academico,
        especialidad_id=datos.especialidad_id,
        centro=centro,
    )

    return {"mensaje": "Nueva rotación solicitada correctamente", "rotacion_id": str(nueva_rotacion.id)}


@router.post("/importar-rotaciones-excel", response_model=schemas.ImportRotacionesResponse)
def importar_rotaciones_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(security.get_current_user),
):
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser XLSX")

    try:
        workbook = load_workbook(BytesIO(file.file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="El Excel no se pudo leer")

    if not workbook.sheetnames:
        raise HTTPException(status_code=400, detail="El Excel no contiene hojas válidas")

    worksheet = workbook[workbook.sheetnames[0]]
    filas = list(worksheet.iter_rows(values_only=True))
    if len(filas) < 2:
        raise HTTPException(status_code=400, detail="El Excel no contiene filas de datos")

    header_map = {}
    for idx, cabecera in enumerate(filas[0]):
        if cabecera is None:
            continue
        clave = _normalizar_texto(cabecera)
        if clave:
            header_map[clave] = idx

    required_headers = {
        "email_alumno": ["email", "correo", "email alumno", "correo alumno"],
        "curso": ["curso", "curso academico"],
        "numero_rotacion": ["rotacion", "numero rotacion", "num rotacion", "rotación"],
        "periodo_academico": ["periodo academico", "año academico", "ano academico"],
        "especialidad": ["especialidad", "unidad"],
        "centro_practicas": ["centro practicas", "centro", "hospital"],
        "email_tutor_hospital": ["email tutor hospital", "tutor hospital", "tutor clinico"],
        "email_tutor_universidad": ["email tutor universidad", "tutor universidad"],
    }

    missing_headers = []
    for campo, aliases in required_headers.items():
        if not any(_normalizar_texto(alias) in header_map for alias in aliases):
            missing_headers.append(campo)
    if missing_headers:
        raise HTTPException(
            status_code=400,
            detail=f"El Excel no contiene las columnas necesarias: {', '.join(missing_headers)}",
        )

    estudiantes = (
        db.query(models.Usuario, models.Alumno)
        .join(models.Alumno, models.Alumno.usuario_id == models.Usuario.id)
        .filter(models.Usuario.rol == "estudiante")
        .all()
    )
    alumnos_por_email = {usuario.email.lower(): alumno for usuario, alumno in estudiantes}

    seen_keys = set()
    total_filas = 0
    creados = []
    fallos = []

    for fila_numero, fila in enumerate(filas[1:], start=2):
        if not any(valor is not None and str(valor).strip() for valor in fila):
            continue

        total_filas += 1
        email_alumno = None
        try:
            email_alumno = str(
                _obtener_valor_requerido(
                    fila,
                    header_map,
                    required_headers["email_alumno"],
                    "email alumno",
                )
            ).strip()
            curso = _extraer_numero(
                _obtener_valor_requerido(fila, header_map, required_headers["curso"], "curso")
            )
            numero_rotacion = _extraer_numero(
                _obtener_valor_requerido(
                    fila,
                    header_map,
                    required_headers["numero_rotacion"],
                    "rotación",
                )
            )
            periodo_academico = str(
                _obtener_valor_requerido(
                    fila,
                    header_map,
                    required_headers["periodo_academico"],
                    "periodo académico",
                )
            ).strip()
            periodo_academico = normalizar_periodo_academico(periodo_academico)
            centro_practicas = str(
                _obtener_valor_requerido(
                    fila,
                    header_map,
                    required_headers["centro_practicas"],
                    "centro de prácticas",
                )
            ).strip()
            email_tutor_hospital = str(
                _obtener_valor_requerido(
                    fila,
                    header_map,
                    required_headers["email_tutor_hospital"],
                    "tutor hospital",
                )
            ).strip()
            email_tutor_universidad = str(
                _obtener_valor_requerido(
                    fila,
                    header_map,
                    required_headers["email_tutor_universidad"],
                    "tutor universidad",
                )
            ).strip()

            if curso not in (2, 3, 4):
                raise ValueError("Curso inválido. Debe ser 2, 3 o 4")
            if numero_rotacion not in (1, 2, 3):
                raise ValueError("Rotación inválida. Debe ser 1, 2 o 3")
            if not periodo_academico:
                raise ValueError("Periodo académico vacío")

            alumno = alumnos_por_email.get(email_alumno.lower())
            if not alumno:
                raise ValueError("No existe un alumno con ese email")

            especialidad = _resolver_especialidad(
                db,
                _obtener_valor_requerido(
                    fila,
                    header_map,
                    required_headers["especialidad"],
                    "especialidad",
                ),
            )
            if not especialidad:
                raise ValueError("La especialidad indicada no existe")

            key = (str(alumno.id), numero_rotacion, curso)
            if key in seen_keys:
                raise ValueError("Duplicado dentro del Excel para el mismo alumno, rotación y curso")

            existe = (
                db.query(models.Rotacion)
                .filter(
                    models.Rotacion.alumno_id == alumno.id,
                    models.Rotacion.numero_rotacion == numero_rotacion,
                    models.Rotacion.curso == curso,
                )
                .first()
            )
            if existe:
                raise ValueError("El alumno ya tiene esa rotación para ese curso")

            rotacion = _crear_rotacion_alumno(
                db=db,
                alumno=alumno,
                curso=curso,
                numero_rotacion=numero_rotacion,
                periodo_academico=periodo_academico,
                centro_practicas=centro_practicas,
                especialidad=especialidad,
                email_tutor_hospital=email_tutor_hospital,
                email_tutor_universidad=email_tutor_universidad,
            )
            seen_keys.add(key)
            creados.append(
                {
                    "fila": fila_numero,
                    "email_alumno": email_alumno,
                    "rotacion_id": str(rotacion.id),
                }
            )
        except HTTPException as exc:
            db.rollback()
            fallos.append(
                {
                    "fila": fila_numero,
                    "email_alumno": email_alumno,
                    "motivo": exc.detail if isinstance(exc.detail, str) else "Error al crear la rotación",
                }
            )
        except ValueError as exc:
            db.rollback()
            fallos.append(
                {
                    "fila": fila_numero,
                    "email_alumno": email_alumno,
                    "motivo": str(exc),
                }
            )
        except Exception:
            db.rollback()
            fallos.append(
                {
                    "fila": fila_numero,
                    "email_alumno": email_alumno,
                    "motivo": "Error inesperado al procesar la fila",
                }
            )

    duplicados = len([f for f in fallos if "duplicado" in f["motivo"].lower()])
    return {
        "total_filas": total_filas,
        "creados": len(creados),
        "duplicados": duplicados,
        "fallidos": len(fallos),
        "creados_detalle": creados,
        "fallos": fallos,
    }


@router.delete("/rotacion/{rotacion_id}")
def eliminar_rotacion(
    rotacion_id: str,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(security.get_current_user),
):
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")

    rotacion = (
        db.query(models.Rotacion).filter(models.Rotacion.id == rotacion_id).first()
    )
    if not rotacion:
        raise HTTPException(status_code=404, detail="Rotación no encontrada")

    db.query(models.RegistroAsistencia).filter(
        models.RegistroAsistencia.rotacion_id == rotacion_id
    ).delete()
    db.query(models.AsignacionTutor).filter(
        models.AsignacionTutor.rotacion_id == rotacion_id
    ).delete()
    db.query(models.CuadernilloRespuesta).filter(
        models.CuadernilloRespuesta.rotacion_id == rotacion_id
    ).delete()

    db.delete(rotacion)
    db.commit()
    return {"mensaje": "Rotación eliminada"}


@router.delete("/{alumno_id}")
def eliminar_alumno_completo(
    alumno_id: str,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(security.get_current_user),
):
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")

    alumno = db.query(models.Alumno).filter(models.Alumno.id == alumno_id).first()
    if not alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")

    usuario_id = alumno.usuario_id
    rotaciones_ids = [r.id for r in alumno.rotaciones]

    db.query(models.RegistroAsistencia).filter(
        or_(
            models.RegistroAsistencia.alumno_id == alumno_id,
            models.RegistroAsistencia.rotacion_id.in_(rotaciones_ids),
        )
    ).delete(synchronize_session=False)
    db.query(models.AsignacionTutor).filter(
        models.AsignacionTutor.rotacion_id.in_(rotaciones_ids)
    ).delete(synchronize_session=False)
    db.query(models.CuadernilloRespuesta).filter(
        models.CuadernilloRespuesta.rotacion_id.in_(rotaciones_ids)
    ).delete(synchronize_session=False)
    db.query(models.Rotacion).filter(models.Rotacion.alumno_id == alumno_id).delete()

    db.delete(alumno)
    usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()
    if usuario:
        db.delete(usuario)

    db.commit()
    return {"mensaje": "Alumno eliminado correctamente"}


@router.get("/asistencia/{rotacion_id}")
def obtener_historial_asistencia_alumno(
    rotacion_id: str,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(security.get_current_user),
):
    if current_user.rol != "estudiante":
        raise HTTPException(status_code=403, detail="Acceso denegado")

    alumno = (
        db.query(models.Alumno)
        .filter(models.Alumno.usuario_id == current_user.id)
        .first()
    )

    fichajes = (
        db.query(models.RegistroAsistencia)
        .filter(
            models.RegistroAsistencia.rotacion_id == rotacion_id,
            models.RegistroAsistencia.alumno_id == alumno.id,
        )
        .order_by(models.RegistroAsistencia.fecha.desc())
        .all()
    )

    # --- ADAPTAMOS PARA DEVOLVER EL MISMO FORMATO QUE EL PROFESOR ---
    registros_limpios = []
    for f in fichajes:
        fecha_str = str(f.fecha).split(" ")[0].split("T")[0] if f.fecha else ""
        email_firmante = f.tutor.email if f.tutor else "Tutor Desconocido"
        fecha_rec_str = str(f.fecha_recuperada).split(" ")[0] if getattr(f, "fecha_recuperada", None) else None

        registros_limpios.append(
            {
                "id": str(f.id),
                "fecha": fecha_str,
                "firmado_en": f.firmado_en.isoformat() if f.firmado_en else "",
                "firmado_por": email_firmante,
                "fecha_recuperada": fecha_rec_str
            }
        )

    return registros_limpios


@router.post("/rotaciones/{rotacion_id}/tutor-campo")
def añadir_tutor_campo(
    rotacion_id: UUID,
    datos: schemas.TutorCampoCreate,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(security.get_current_user),
):
    if current_user.rol != "estudiante":
        raise HTTPException(status_code=403, detail="Solo los alumnos pueden añadir un tutor de campo")

    alumno = db.query(models.Alumno).filter(models.Alumno.usuario_id == current_user.id).first()
    if not alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")

    rotacion = db.query(models.Rotacion).filter(
        models.Rotacion.id == rotacion_id,
        models.Rotacion.alumno_id == alumno.id
    ).first()
    
    if not rotacion:
        raise HTTPException(status_code=404, detail="Rotación no encontrada o no pertenece al alumno")

    if rotacion.completada:
        raise HTTPException(status_code=400, detail="No se puede añadir tutor a una rotación completada")

    # Comprobar si ya existe una asignación de tutor de campo
    asignacion_existente = db.query(models.AsignacionTutor).filter(
        models.AsignacionTutor.rotacion_id == rotacion.id,
        models.AsignacionTutor.tipo_tutor == "campo"
    ).first()

    if asignacion_existente:
        raise HTTPException(status_code=400, detail="Ya existe un tutor de campo asignado a esta rotación")

    # Buscar usuario existente o crearlo
    tutor_usuario = db.query(models.Usuario).filter(models.Usuario.email == datos.email).first()
    
    es_nuevo = False
    if not tutor_usuario:
        es_nuevo = True
        import secrets
        tutor_usuario = models.Usuario(
            email=datos.email,
            password_hash=security.get_password_hash(secrets.token_urlsafe(16)),
            rol="profesor",
            tipo_tutor="campo",
            registro_completado=False,
            activo=True
        )
        db.add(tutor_usuario)
        db.flush()

    # Asignar a la rotación
    nueva_asignacion = models.AsignacionTutor(
        tutor_id=tutor_usuario.id,
        rotacion_id=rotacion.id,
        tipo_tutor="campo"
    )
    db.add(nueva_asignacion)
    db.flush()

    alumno_nombre = f"{security.descifrar_dato(alumno.nombre_cifrado)} {security.descifrar_dato(alumno.apellidos_cifrado)}"
    especialidad_nombre = rotacion.especialidad.nombre if rotacion.especialidad else "Sin especialidad"

    # Enviar invitación al tutor de campo
    if es_nuevo:
        import secrets
        from datetime import datetime, timezone, timedelta
        import os
        token_hex = secrets.token_urlsafe(32)
        nuevo_token = models.TokenRecuperacion(
            usuario_id=tutor_usuario.id,
            token=token_hex,
            expira_at=datetime.now(timezone.utc) + timedelta(minutes=30),
        )
        db.add(nuevo_token)
        
        frontend_url = os.getenv("FRONTEND_URL", "http://localhost:3000")
        enlace = f"{frontend_url}/restablecer-password?token={token_hex}"
        
        from ..utils.email_utils import enviar_invitacion_tutor_campo
        enviar_invitacion_tutor_campo(datos.email, alumno_nombre, especialidad_nombre, enlace)
    else:
        from ..utils.email_utils import enviar_aviso_asignacion_tutor_existente
        enviar_aviso_asignacion_tutor_existente(datos.email, alumno_nombre, especialidad_nombre)

    db.commit()

    # Notificar a los otros tutores
    from ..utils.email_utils import enviar_aviso_nuevo_tutor_campo
    tutores_emails = []
    asignaciones_existentes = db.query(models.AsignacionTutor).filter(
        models.AsignacionTutor.rotacion_id == rotacion.id,
        models.AsignacionTutor.tipo_tutor.in_(["hospital", "universidad"])
    ).all()
    for asig in asignaciones_existentes:
        tutores_emails.append(asig.tutor.email)

    if tutores_emails:
        enviar_aviso_nuevo_tutor_campo(tutores_emails, alumno_nombre, especialidad_nombre, datos.email)

    return {"mensaje": "Tutor de campo añadido correctamente"}
